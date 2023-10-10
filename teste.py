import PyPDF2
import tkinter as tk
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import re


def select_pdf():
    root = tk.Tk()
    root.withdraw()  # Ocultar a janela principal
    file_path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
    return file_path


def process_qualitplacas():
    file_path = select_pdf()

    if not file_path:
        print("Nenhum arquivo PDF selecionado.")
        return

    # Abrir o arquivo PDF selecionado
    with open(file_path, "rb") as pdf_file:
        dados_pdf = PyPDF2.PdfReader(pdf_file)
        linhas_imprimir = True  # Inicializado como True para imprimir a primeira página
        linhas_impressas = 0  # Dicionário para armazenar os valores por dia
        data_list = []
        descricao_list = []
        valor_list = []

        for pagina_num, pagina in enumerate(dados_pdf.pages, 1):
            linhas_quebra = 0  # Inicialize aqui

            if pagina_num == 1:
                # Apenas para a primeira página
                texto_pagina = pagina.extract_text()
                linhas = texto_pagina.split("\n")

                for linha in linhas:
                    if "Saldo da conta" in linha:
                        linhas_imprimir = False
                        break  # Parar de imprimir quando encontrar "Saldo da conta"

                    if linhas_imprimir and linhas_quebra >= 6:
                        partes = linha.split()
                        if len(partes) > 5:
                            data = partes[0]
                            valor = partes[-2]
                            descricao = " ".join(partes[1:-4])

                            data_list.append(data)
                            descricao_list.append(descricao)
                            valor_list.append(valor)

                            # Configurar a cor da fonte com base em "-" em valor
                            font_color = "blue" if "-" not in valor else "red"

                            # Realizar a formatação das informações nas linhas
                            linha_formatada = (
                                f"Data: {data}  Descrição: {descricao}  Valor: {valor}"
                            )

                            # Imprimir a linha formatada com a cor da fonte

                    linhas_impressas += 1
                    linhas_quebra += 1
            else:
                # Nas outras páginas, imprimir normalmente se linhas_imprimir for True
                if linhas_imprimir:
                    texto_pagina = pagina.extract_text()
                    linhas = texto_pagina.split("\n")

                    for linha in linhas:
                        if "Saldo da conta" in linha:
                            linhas_imprimir = False
                            break  # Parar de imprimir quando encontrar "Saldo da conta"

                        partes = linha.split()
                        if len(partes) > 5:
                            data = partes[0]
                            valor = partes[-2]
                            descricao = " ".join(partes[1:-4])

                            data_list.append(data)
                            descricao_list.append(descricao)
                            valor_list.append(valor)

                            # Configurar a cor da fonte com base em "-" em valor
                            font_color = "blue" if "-" not in valor else "red"

                            # Realizar a formatação das informações nas linhas
                            linha_formatada = (
                                f"Data: {data}  Descrição: {descricao}  Valor: {valor}"
                            )

                            # Imprimir a linha formatada com a cor da fonte

                    linhas_impressas += 1

    # Criar um DataFrame com os dados
    df = pd.DataFrame(
        {"DATA": data_list, "DESCRICAO": descricao_list, "VALOR": valor_list}
    )

    # Substituir "." por "" e "," por "."
    df["VALOR"] = df["VALOR"].str.replace(".", "").str.replace(",", ".")

    # Solicitar o local para salvar o arquivo Excel
    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")]
    )
    if save_path:
        # Salvar o DataFrame como um arquivo Excel
        df.to_excel(save_path, index=False)

        # Carregar o arquivo Excel
        wb = load_workbook(filename=save_path)
        ws = wb.active

        # Configurar estilos para as colunas de VALOR
        red_font = Font(color="FF0000")
        blue_font = Font(color="0000FF")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
            for cell in row:
                if "-" in cell.value:
                    cell.font = red_font
                    cell.value = cell.value.replace("-", "")
                else:
                    cell.font = blue_font

        # Salvar as alterações
        wb.save(save_path)
        print(f"Arquivo Excel salvo em: {save_path}")


if __name__ == "__main__":
    process_qualitplacas()
